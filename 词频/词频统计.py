# 导入依赖
import os,psutil,time
from openpyxl import Workbook
from openpyxl import load_workbook
import multiprocessing
current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path))
#import jieba
# 获取关键词列表
#os.chdir(sys.path[0])
def getKeyWordList():
    # 加载 Excel 文件
    workbook = load_workbook(keyword_dir)
    # 选择第一个工作表
    sheet = workbook.active
    return [cell.value for cell in sheet['A']]

def getKeyWordData(text_path,file,key_word):
    # 读取文本
    data = []
    path_filename = text_path + file
    with  open(path_filename, "r", encoding='utf-8') as f:
        txt = f.read()
        # 使用精确模式对文本进行分词
        #words = jieba.lcut(txt)
        year = file[7:11]#获取年份
        code = file[:6]#获取公司代码
        name = ''
        for chr in file[::-1][4:]: 
            if chr == "-":
                break
            name = name + chr
        
        data.append(code)
        data.append(name[::-1])
        data.append(year) 

        for wd in key_word:
            data.append(txt.count(wd))#统计关键词出现次数
            
    return data  
def statistics(num,file_lst,key_word,lock):
    datas = []
    for root ,file in file_lst:
        if file.endswith(".txt"):
            data = getKeyWordData( root,file,key_word)
            #os.remove("年度报告/"+folder_name+"/"+file)
            datas.append(data)
        else:
            print(file+"不是txt文件")
            continue#return改为continue，直接去掉return也可以。
   
    
    lock.acquire()
    book = load_workbook("词频统计.xlsx")
    sheet = book.active
    for row in datas:
        sheet.append(row)

    book.save("词频统计.xlsx")
    book.close()
    lock.release()
    print(f"第{num}组统计完成")
    
# 主函数
def main():
    key_word = getKeyWordList()
    lst = ["企业代码","企业名称","年份"]
    lst.extend(key_word)
    book = Workbook()
    sheet = book.active
    sheet.append(lst)
    
    """
    book = Workbook()
    book.save("词频统计.xlsx")
    上面两行代码会新建一个文件，且会覆盖已有同名文件
    下面第一行代码则是导入已存在的文件，不存在则报错
    book = load_workbook("词频统计.xlsx")
    book.save("词频统计.xlsx")

    """
    book.save("词频统计.xlsx")
    book.close()

    file_dir_lst = []
    lock = multiprocessing.Manager().Lock()
    
    for root, dirs, files in os.walk(base_dir):
        file_dir_lst.extend([(root + "\\",file) for file in files])
    
    pool = multiprocessing.Pool(processes = psutil.cpu_count()+1)#使用多进程，提高统计速度
    group_count = 2
    total_group = len(file_dir_lst) // group_count
    print(f"共有{len(file_dir_lst)}个文件，分为{total_group+1}组,每组{group_count}，最后一组为余数，从第0组")
    for  num in range(total_group):
        pool.apply_async(statistics, (num,file_dir_lst[num* group_count:(num+1)*group_count],key_word,lock) )#

    pool.apply_async(statistics,(num+1,file_dir_lst[total_group*(group_count):],key_word,lock))#
    

    pool.close()
    pool.join()
    

base_dir = "2022测试"
keyword_dir =  "关键词.xlsx"
cipin_dir =  "词频统计.csv"


def countdown(seconds):
    for i in range(seconds, 0, -1):
        print(i, end='\r')
        time.sleep(1)


if __name__ == '__main__':
    print("请勿强制退出，否则导致数据损坏")
    countdown(5)
    main()
   
   


